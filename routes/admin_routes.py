import os
import time
from flask import Blueprint, render_template, jsonify, request
from werkzeug.utils import secure_filename
from models import db
# 스키마 임포트
from schema.customer_schema import Reservation
from schema.chat_schema import ChatRoom, ChatLog
from schema.land_schema import LandChatRoom, LandChatLog
from schema.schema import ProductTable
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# [핵심 수정] url_prefix 제거! (app.py에서 /admin으로 등록됨)
bp = Blueprint('admin', __name__, template_folder='../templates')

# 파일 업로드 설정
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'json', 'txt', 'xlsx'}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# =========================================================
# 1. 페이지 라우트
# =========================================================

@bp.route('/')
def dashboard():
    return render_template('admin/index.html', active_page='dashboard')


@bp.route('/products')
def product_list():
    # [참고] 추후 DB 연동 시 ProductTable.query.all() 등으로 변경 가능
    products = [
        {'id': 'P001', 'name': '도쿄 3박 4일 패키지', 'status': '판매중', 'period': '3박 4일', 'manager': '김관리', 'region': '일본'},
        {'id': 'P002', 'name': '오사카 유니버설 스튜디오', 'status': '대기', 'period': '2박 3일', 'manager': '이매니저', 'region': '일본'},
        {'id': 'P003', 'name': '다낭 호캉스 특가', 'status': '판매중', 'period': '3박 5일', 'manager': '박팀장', 'region': '베트남'},
        {'id': 'P004', 'name': '제주도 힐링 여행', 'status': '종료', 'period': '2박 3일', 'manager': '최대리', 'region': '한국'},
    ]
    return render_template('admin/product_list.html', active_page='products', products=products)

@bp.route('/products/new')
def product_create():
    return render_template('admin/product_create.html', active_page='products')

@bp.route('/products/<id>')
def product_detail(id):
    product = {'id': id, 'name': '도쿄 3박 4일 패키지', 'status': '판매중', 'period': '3박 4일', 'manager': '김관리', 'region': '일본',
               'price': '1,200,000'}
    return render_template('admin/product_detail.html', active_page='products', product=product)

@bp.route('/reservations')
def reservation_list():
    reservations = [
        {'id': 'RES-2025-001', 'customer': '홍길동', 'phone': '010-1234-5678', 'product': '미야코지마 골프 3박 4일',
         'date': '2025-10-10', 'pax': 2, 'status': '예약확정'},
        {'id': 'RES-2025-002', 'customer': '김철수', 'phone': '010-2345-6789', 'product': '도쿄 3일 패키지',
         'date': '2025-10-15', 'pax': 4, 'status': '대기'},
        {'id': 'RES-2025-003', 'customer': '이영희', 'phone': '010-3456-7890', 'product': '오사카 4일 패키지',
         'date': '2025-10-20', 'pax': 2, 'status': '예약확정'},
        {'id': 'RES-2025-004', 'customer': '박민수', 'phone': '010-4567-8901', 'product': '제주도 3박 4일',
         'date': '2025-10-25', 'pax': 3, 'status': '취소'},
        {'id': 'RES-2025-005', 'customer': '정수진', 'phone': '010-5678-9012', 'product': '부산 2박 3일', 'date': '2025-11-01',
         'pax': 2, 'status': '대기'},
    ]
    return render_template('admin/reservation_list.html', active_page='reservations', reservations=reservations)

@bp.route('/reservations/<id>')
def reservation_detail(id):
    reservation = {'id': id, 'customer': '홍길동', 'phone': '010-1234-5678', 'product': '미야코지마 골프', 'date': '2025-10-10',
                   'pax': 2, 'status': '예약확정', 'amount': '2,500,000'}
    return render_template('admin/reservation_detail.html', active_page='reservations', reservation=reservation)

@bp.route('/quotations')
def quotation_list():
    quotations = [
        {'id': 'Q-2025-001', 'customer': '홍길동', 'product': '미야코지마 골프 3박 4일', 'amount': '2,500,000',
         'date': '2025-01-15', 'status': '대기'},
        {'id': 'Q-2025-002', 'customer': '김철수', 'product': '도쿄 3일 패키지', 'amount': '1,800,000', 'date': '2025-01-16',
         'status': '확정'},
        {'id': 'Q-2025-003', 'customer': '이영희', 'product': '오사카 4일 패키지', 'amount': '2,200,000', 'date': '2025-01-17',
         'status': '대기'},
        {'id': 'Q-2025-004', 'customer': '박민수', 'product': '제주도 3박 4일', 'amount': '1,500,000', 'date': '2025-01-18',
         'status': '취소'},
        {'id': 'Q-2025-005', 'customer': '정수진', 'product': '부산 2박 3일', 'amount': '800,000', 'date': '2025-01-19',
         'status': '확정'},
    ]
    return render_template('admin/quotation_list.html', active_page='quotations', quotations=quotations)

@bp.route('/quotations/new')
def quotation_create():
    return render_template('admin/quotation_create.html', active_page='quotations')

@bp.route('/quotations/<id>')
def quotation_detail(id):
    quotation = {'id': id, 'customer': '홍길동', 'product': '미야코지마 골프', 'amount': '2,500,000', 'date': '2025-01-15',
                 'status': '대기'}
    return render_template('admin/quotation_detail.html', active_page='quotations', quotation=quotation)

@bp.route('/payments')
def payment_page():
    return render_template('admin/payment.html', active_page='payments')

@bp.route('/finance')
def finance_page():
    return render_template('admin/finance.html', active_page='finance')

@bp.route('/flights')
def flight_list():
    flights = [
        {'id': 'KE001', 'airline': '대한항공', 'route': 'ICN → NRT', 'departure': '2025-10-10 09:00',
         'arrival': '2025-10-10 12:00', 'pax': 2, 'status': '예약완료'},
        {'id': 'OZ201', 'airline': '아시아나항공', 'route': 'ICN → KIX', 'departure': '2025-10-15 14:00',
         'arrival': '2025-10-15 16:30', 'pax': 4, 'status': '발권대기'},
        {'id': '7C123', 'airline': '제주항공', 'route': 'ICN → CJU', 'departure': '2025-10-20 08:00',
         'arrival': '2025-10-20 09:30', 'pax': 2, 'status': '예약완료'},
        {'id': 'LJ501', 'airline': '진에어', 'route': 'ICN → PUS', 'departure': '2025-10-25 10:00',
         'arrival': '2025-10-25 11:00', 'pax': 3, 'status': '발권완료'},
        {'id': 'TW301', 'airline': '티웨이항공', 'route': 'ICN → FUK', 'departure': '2025-11-01 13:00',
         'arrival': '2025-11-01 15:00', 'pax': 2, 'status': '예약완료'},
    ]
    return render_template('admin/flight_list.html', active_page='flights', flights=flights)

@bp.route('/hotels')
def hotel_list():
    hotels = [
        {'name': '미야코지마 리조트', 'region': '오키나와, 일본', 'rating': 5, 'rooms': 120, 'manager': '김호텔'},
        {'name': '도쿄 그랜드 호텔', 'region': '도쿄, 일본', 'rating': 4, 'rooms': 200, 'manager': '이호텔'},
        {'name': '오사카 베이 호텔', 'region': '오사카, 일본', 'rating': 4, 'rooms': 150, 'manager': '박호텔'},
        {'name': '제주 신라호텔', 'region': '제주도, 한국', 'rating': 5, 'rooms': 300, 'manager': '정호텔'},
        {'name': '부산 파라다이스 호텔', 'region': '부산, 한국', 'rating': 5, 'rooms': 250, 'manager': '최호텔'},
    ]
    return render_template('admin/hotel_list.html', active_page='hotels', hotels=hotels)

@bp.route('/attractions')
def attraction_list():
    attractions = [
        {'name': '도쿄 타워', 'region': '도쿄, 일본', 'type': '랜드마크', 'time': '2시간'},
        {'name': '오사카 성', 'region': '오사카, 일본', 'type': '역사', 'time': '1.5시간'},
        {'name': '제주 한라산', 'region': '제주도, 한국', 'type': '자연', 'time': '4시간'},
        {'name': '부산 해운대', 'region': '부산, 한국', 'type': '해변', 'time': '3시간'},
        {'name': '교토 기요미즈데라', 'region': '교토, 일본', 'type': '사원', 'time': '2시간'},
    ]
    return render_template('admin/attraction_list.html', active_page='attractions', attractions=attractions)

@bp.route('/partners')
def partner_list():
    partners = [
        {'id': 1, 'name': '일본여행사', 'region': '도쿄, 일본', 'type': '랜드오퍼레이터', 'contact': '김파트너', 'phone': '010-1111-2222',
         'email': 'kim@japan-travel.com', 'status': 'active'},
        {'id': 2, 'name': '제주여행사', 'region': '제주도, 한국', 'type': '로컬여행사', 'contact': '이파트너', 'phone': '010-2222-3333',
         'email': 'lee@jeju-travel.com', 'status': 'active'},
        {'id': 3, 'name': '부산여행사', 'region': '부산, 한국', 'type': '로컬여행사', 'contact': '박파트너', 'phone': '010-3333-4444',
         'email': 'park@busan-travel.com', 'status': 'active'},
        {'id': 4, 'name': '오키나와여행사', 'region': '오키나와, 일본', 'type': '랜드오퍼레이터', 'contact': '정파트너',
         'phone': '010-4444-5555', 'email': 'jung@okinawa-travel.com', 'status': 'active'},
        {'id': 5, 'name': '도쿄여행사', 'region': '도쿄, 일본', 'type': '랜드오퍼레이터', 'contact': '최파트너', 'phone': '010-5555-6666',
         'email': 'choi@tokyo-travel.com', 'status': 'inactive'},
    ]
    return render_template('admin/partner_list.html', active_page='partners', partners=partners)

@bp.route('/partners/new')
def partner_create():
    return render_template('admin/partner_create.html', active_page='partners')

@bp.route('/customers')
def customer_list():
    customers = [
        {'id': 'C-001', 'name': '홍길동', 'email': 'hong@example.com', 'phone': '010-1234-5678', 'trips': 3,
         'lastTrip': '2025-10-10', 'grade': 'VIP', 'status': 'active'},
        {'id': 'C-002', 'name': '김철수', 'email': 'kim@example.com', 'phone': '010-2345-6789', 'trips': 2,
         'lastTrip': '2025-10-15', 'grade': 'Gold', 'status': 'active'},
        {'id': 'C-003', 'name': '이영희', 'email': 'lee@example.com', 'phone': '010-3456-7890', 'trips': 1,
         'lastTrip': '2025-10-20', 'grade': '일반', 'status': 'active'},
        {'id': 'C-004', 'name': '박민수', 'email': 'park@example.com', 'phone': '010-4567-8901', 'trips': 0,
         'lastTrip': '-', 'grade': '일반', 'status': 'inactive'},
        {'id': 'C-005', 'name': '정수진', 'email': 'jung@example.com', 'phone': '010-5678-9012', 'trips': 1,
         'lastTrip': '2025-11-01', 'grade': '일반', 'status': 'active'},
    ]
    return render_template('admin/customer_list.html', active_page='customers', customers=customers)

@bp.route('/settings')
def settings_page():
    return render_template('admin/settings.html', active_page='settings')

# =========================================================
# 2. 채팅 API (고객용) -> 최종 주소: /admin/rooms
# =========================================================

@bp.route('/rooms')
def get_rooms():
    try:
        rooms = ChatRoom.query.order_by(ChatRoom.last_active.desc()).all()
        result = []
        for r in rooms:
            result.append({
                'session_id': r.session_id,
                'user_name': r.user_name,
                'last_message': r.last_message,
                'last_active': r.last_active.strftime('%Y-%m-%d %H:%M:%S') if r.last_active else '',
                'status': r.status
            })
        return jsonify(result)
    except Exception as e:
        print(f"Error fetching rooms: {e}")
        return jsonify([])


@bp.route('/history/<session_id>')
def get_history(session_id):
    try:
        logs = ChatLog.query.filter_by(session_id=session_id) \
            .order_by(ChatLog.created_at.asc()).all()
        return jsonify([l.to_dict() for l in logs])
    except Exception as e:
        print(f"Error fetching history: {e}")
        return jsonify([])


# =========================================================
# 3. 채팅 API (랜드사용) -> 최종 주소: /admin/land-rooms
# =========================================================

@bp.route('/land-rooms')
def get_land_rooms():
    try:
        rooms = LandChatRoom.query.order_by(LandChatRoom.last_active.desc()).all()
        result = []
        for r in rooms:
            result.append({
                'session_id': r.session_id,
                'operator_name': r.operator_name,
                'last_message': r.last_message,
                'last_active': r.last_active.strftime('%Y-%m-%d %H:%M:%S') if r.last_active else '',
                'status': r.status
            })
        return jsonify(result)
    except Exception as e:
        print(f"Error fetching land rooms: {e}")
        return jsonify([])


@bp.route('/land-history/<session_id>')
def get_land_history(session_id):
    try:
        logs = LandChatLog.query.filter_by(session_id=session_id) \
            .order_by(LandChatLog.created_at.asc()).all()
        return jsonify([l.to_dict() for l in logs])
    except Exception as e:
        print(f"Error fetching land history: {e}")
        return jsonify([])


# =========================================================
# 4. 공통 기능
# =========================================================

@bp.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        import time
        filename = f"{int(time.time())}_{filename}"

        save_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(save_path)

        return jsonify({
            'success': True,
            'filename': filename,
            'filepath': f"/static/uploads/{filename}"
        })

    return jsonify({'error': 'Not allowed type'}), 400


@bp.route('/generate-inquiry/<int:reservation_id>')
def generate_inquiry(reservation_id):
    try:
        res = Reservation.query.get(reservation_id)
        if not res:
            return jsonify({'error': 'Reservation not found'}), 404

        product_name = "상품 정보 없음"
        country = "해외"
        if res.product_id:
            prod = ProductTable.query.get(res.product_id)
            if prod:
                product_name = f"{prod.product_id}번, {prod.product_name}"
                country = prod.country

        template = f"""-[{country}] 예약문의-

대표자 성함 : {res.rep_name}
출발지 : {res.departure_place or '미정'}
인원 : {res.headcount}명
출발일 : {res.start_date if hasattr(res, 'start_date') else '일정표 참조'}
요청 사항 : {res.requests}

상품정보 : {product_name}
"""
        return jsonify({'success': True, 'text': template})

    except Exception as e:
        print(f"Generate inquiry error: {e}")
        return jsonify({'error': str(e)}), 500

# =========================================================
# [NEW] 템플릿 없이 코드로 엑셀 인보이스 직접 그리기
# =========================================================
@bp.route('/generate-invoice', methods=['POST'])
def generate_invoice():
    try:
        # 1. 데이터 수신 및 파싱
        data = request.get_json()
        if not data: return jsonify({'error': 'No data provided'}), 400
        if isinstance(data, list): data = data[0]

        # 데이터 정리
        customer_name = data.get('customer_name', '홍길동 고객님')
        product_name = data.get('product_name', '여행 상품')

        start_date = data.get('start_date', '')
        end_date = data.get('end_date', '')
        nights = data.get('nights', 0)
        days = data.get('days', 0)
        period = f"{start_date} ~ {end_date} ({nights}박 {days}일)"

        price_adult = data.get('price_adult', 0)
        head_counts = data.get('head_counts', 1)
        total_price = data.get('total_price', price_adult * head_counts)

        # 상세 내용 파싱
        details_raw = data.get('details', '{}')
        if isinstance(details_raw, str):
            try:
                details = json.loads(details_raw)
            except:
                details = {}
        else:
            details = details_raw

        inclusions = "\n".join([f"• {item}" for item in details.get('inclusions', [])])
        exclusions = "\n".join([f"• {item}" for item in details.get('exclusions', [])])

        # ---------------------------------------------------------
        # 2. 엑셀 그리기 시작 (openpyxl)
        # ---------------------------------------------------------
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"

        # (1) 스타일 정의
        font_title = Font(name='맑은 고딕', size=24, bold=True, color='1F4E78')
        font_header = Font(name='맑은 고딕', size=12, bold=True, color='FFFFFF')
        font_text = Font(name='맑은 고딕', size=11)
        font_bold = Font(name='맑은 고딕', size=11, bold=True)

        fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')  # 짙은 파랑
        fill_light = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')  # 연한 하늘색

        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center')

        # (2) 열 너비 조정
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 40  # 상품명
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20  # 단가
        ws.column_dimensions['F'].width = 10  # 인원
        ws.column_dimensions['G'].width = 20  # 합계

        # (3) 타이틀 작성
        ws.merge_cells('B2:G3')
        title_cell = ws['B2']
        title_cell.value = "INVOICE"
        title_cell.font = font_title
        title_cell.alignment = align_center

        # (4) 기본 정보 (수신자, 날짜)
        ws['B5'] = "To:"
        ws['B5'].font = font_bold
        ws['C5'] = customer_name
        ws['C5'].font = font_text

        ws['F5'] = "Date:"
        ws['F5'].font = font_bold
        ws['F5'].alignment = align_right
        ws['G5'] = time.strftime('%Y-%m-%d')
        ws['G5'].font = font_text
        ws['G5'].alignment = align_right

        # (5) 상품 테이블 헤더
        headers = [('B7', 'Description'), ('E7', 'Unit Price'), ('F7', 'Qty'), ('G7', 'Amount')]
        ws.merge_cells('B7:D7')  # 상품명 칸 병합

        for cell_loc, text in headers:
            cell = ws[cell_loc]
            cell.value = text
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin

        # 병합된 헤더 테두리 처리
        for col in range(2, 5): ws.cell(row=7, column=col).border = border_thin

        # (6) 상품 데이터 입력
        ws.merge_cells('B8:D8')

        # 내용
        ws['B8'] = f"{product_name}\n({period})"
        ws['E8'] = price_adult
        ws['F8'] = head_counts
        ws['G8'] = total_price

        # 스타일 적용
        for col in ['B', 'E', 'F', 'G']:
            cell = ws[f'{col}8']
            cell.font = font_text
            cell.border = border_thin
            cell.alignment = align_center
            if col == 'B': cell.alignment = align_left
            if col in ['E', 'G']: cell.number_format = '#,##0'  # 천단위 콤마

        # (7) 총계 (Total)
        ws['F10'] = "TOTAL"
        ws['F10'].font = font_bold
        ws['F10'].alignment = align_right

        ws['G10'] = total_price
        ws['G10'].font = font_bold
        ws['G10'].number_format = '#,##0'
        ws['G10'].fill = fill_light
        ws['G10'].border = border_thin
        ws['G10'].alignment = align_center

        # (8) 상세 내용 (포함/불포함)
        # 헤더
        ws.merge_cells('B12:D12')
        ws.merge_cells('E12:G12')

        ws['B12'] = "Inclusions (포함 사항)"
        ws['E12'] = "Exclusions (불포함 사항)"

        for cell in [ws['B12'], ws['E12']]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin

        # 내용
        ws.merge_cells('B13:D20')  # 포함사항 박스 크게
        ws.merge_cells('E13:G20')  # 불포함사항 박스 크게

        ws['B13'] = inclusions
        ws['E13'] = exclusions

        for cell in [ws['B13'], ws['E13']]:
            cell.font = font_text
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = border_thin

        # 테두리 보정 (병합된 영역)
        for r in range(13, 21):
            ws.cell(row=r, column=4).border = border_thin  # D열 우측
            ws.cell(row=r, column=7).border = border_thin  # G열 우측
        for c in range(2, 5): ws.cell(row=20, column=c).border = border_thin  # 20행 하단 (왼쪽)
        for c in range(5, 8): ws.cell(row=20, column=c).border = border_thin  # 20행 하단 (오른쪽)

        # 3. 파일 저장
        timestamp = int(time.time())
        filename = f"Invoice_{timestamp}.xlsx"
        save_path = os.path.join(UPLOAD_FOLDER, filename)

        wb.save(save_path)
        wb.close()

        return jsonify({
            'success': True,
            'message': '인보이스 생성 완료',
            'filename': filename,
            'filepath': f"/static/uploads/{filename}",
            'download_url': f"/static/uploads/{filename}"
        })

    except Exception as e:
        print(f"❌ Invoice Generation Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500